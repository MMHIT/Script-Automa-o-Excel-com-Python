# Bot Financeiro inserção de dados.
   
from datetime       import date
from openpyxl       import Workbook
from openpyxl.utils import get_column_letter
from openpyxl       import load_workbook
from openpyxl       import Workbook
import o
import json

# Data Atualizada
    data_atual = date.today()
    data_em_texto = data_atual.strftime('%d-%m-%Y')

#Disparo de e-mail ao Banco
 
# Edita as informações e anexa os arquivos
msg = EmailMessage()
msg['Subject'] = config_file.email.assunto
msg['From']    = config_file.user.username
msg['To']      = config_file.email.destinatarios_email
msg['Cc']      = config_file.email.destinatarios_email_cc
msg['Bcc']     = config_file.email.destinatarios_email_cco
msg.set_content(config_file.email.mensagem + config_file.email.assinatura)
#adicionar_anexos(msg, test)
# msg.add_attachment()
#
# envia a mensagem
with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
    smtp.login(config_file.user.username, config_file.user.password)
    smtp.send_message(msg)

# Envia o e-mail

except Exception as e:
     print('\n\n\n\nErro!!!!\n')
     print(e)
     input('\nDigite enter para finalizar.')
     raise


#Pegando o arquivo do banco



#lendo o arquivo do banco

wb = load_workbook('Entradas_3298.xlsx') # abrindo o Workbook test.xlsx
ws = wb['Page 1'] # selecionando a planilha 'Plan1' dentro do Workbook test.xlsx

for line in ws:  # iterando em todas as linhas da 'Plan1'
    print line[3] # print a primeira célula da linha


wb = load_workbook('Liquidacoes_3298.xlsx') # abrindo o Workbook test.xlsx
ws = wb['Page 1'] # selecionando a planilha 'Plan1' dentro do Workbook test.xlsx

for line in ws:  # iterando em todas as linhas da 'Plan1'
    print line[3] # print a primeira célula da linha


wb = load_workbook('Entradas_15423.xlsx') # abrindo o Workbook test.xlsx
ws = wb['Page 1'] # selecionando a planilha 'Plan1' dentro do Workbook test.xlsx

for line in ws:  # iterando em todas as linhas da 'Plan1'
    print line[3] # print a primeira célula da linha
                    

wb = load_workbook('Liquidacoes_15423.xlsx') # abrindo o Workbook test.xlsx
ws = wb['Page 1'] # selecionando a planilha 'Plan1' dentro do Workbook test.xlsx

for line in ws:  # iterando em todas as linhas da 'Plan1'
    print line[3] # print a primeira célula da linha



#Processando Arquivo  
 
#Cria diretório
os.makedirs(data_em_texto + "_Processado")            
        
#Cria uma pasta de trabalho no excel(arquivo) em excel
Goiania_entrada = Workbook()
#Ativa a planilha no excel
planilha1 = Goiania_entrada.active
#Nomeia a planilha criada
planilha1.title = "3298_entrada_" + data_em_texto
Goiania_entrada.save("3298_entrada_" + data_em_texto + ".xlsx") #Salva arquivo com nome e data referente

Goiania_liquidação = Workbook()
#Ativa a planilha no excel
planilha1 = Goiania_liquidação.active
#Nomeia a planilha criada
planilha1.title = "3298_liquid_" + data_em_texto
Goiania_liquidação.save("3298_liquid_" + data_em_texto + ".xlsx") #Salva arquivo com nome e data referente

#Cria uma pasta de trabalho no excel(arquivo) em excel
Aparecida_entrada = Workbook()
#Ativa a planilha no excel
planilha1 = Aparecida_entrada.active
#Nomeia a planilha criada
planilha1.title = "15423_entrada_" + data_em_texto
Aparecida_entrada.save("15423_entrada_" + data_em_texto + ".xlsx") #Salva arquivo com nome e data referente

Aparecida_liquidação = Workbook()
#Ativa a planilha no excel
planilha1 = Aparecida_liquidação.active
#Nomeia a planilha criada
planilha1.title = "15423_liquid_" + data_em_texto
Aparecida_liquidação.save("15423_liquid_" + data_em_texto + ".xlsx") #Salva arquivo com nome e data referente



    #Copiando arquivos do Excel sicoob para o novo arquivo
    original = arquivo_excel.get_sheet_by_name('Entradas_3298.xlsx')
    copia = arquivo_excel.copy_worksheet(copia)
    arquivo_excel.save("3298_entrada_" + data_em_texto +".xlsx")

    #Copiando arquivos do Escel sicoob para o novo arquivo
    original = arquivo_excel.get_sheet_by_name('Liquidacoes_3298.xlsx')
    copia = arquivo_excel.copy_worksheet(copia)
    arquivo_excel.save("3298_liquid_" + data_em_texto + ".xlsx")


#2º exemplo de nomeação (******planilha2 = Goiania_entrada.create_sheet("Nova entrada_27-05-29_05_2020_3")*********)
#3º exemplo de nomeação em posição definida (****planilha2 = Goiania_liquidação.create_sheet("Nova entrada_27-05-29_05_2020_3", 0)*****)
    
# Usando arquivos VBA
import os, os.path
import win32com.client

if os.path.exists('C:\Users\pvmatheus\Desktop\BotExcelSicoob\Macros\copia_sicoob.bas'):
    Excel_macro = win32com.client.DispatchEx("Excel.Application") # DispatchEx is required in the newest versions of Python.
    Excel_path = os.path.expanduser('C:\Users\pvmatheus\Desktop\BotExcelSicoob\Macros\copia_sicoob.bas')
    workbook = Excel_macro.Workbooks.Open(Filename = Excel_path, ReadOnly =1)
    Excel_macro.Application.Run("copia_sicoob.bas!Module1.Macro1") # update Module1 with your module, Macro1 with your macro
    workbook.Save()
    Excel_macro.Application.Quit()  
    del Excel_macro

#Formatando o arquivo
wb = Workbook()

dest_filename = '3298_entrada_.xlsx'
            
ws1 = wb.active
ws1.title = "3298_entrada_"
            
for row in range(1, 40):
    ws1.append(range(600))
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
    _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)
    AA
    wb.save(filename = dest_filename)

#Salva em CSV
       